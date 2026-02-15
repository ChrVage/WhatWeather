"""
Excel Formatter - Convert API response data to Excel format
"""

from datetime import datetime
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelFormatter:
    """Format API responses as Excel spreadsheets"""
    
    @staticmethod
    def format(data, filename, title="API Response"):
        """
        Convert data to Excel format and save to file
        
        Args:
            data: Dictionary containing API response
            filename: Path to save Excel file
            title: Title for the spreadsheet
        """
        wb = Workbook()
        
        # Check if this is timeseries data (MET Norway format)
        if ExcelFormatter._is_timeseries_data(data):
            ExcelFormatter._format_timeseries(data, wb, title)
        else:
            # Use generic format for non-timeseries data
            ExcelFormatter._format_generic(data, wb, title)
        
        # Save file
        wb.save(filename)
        return filename
    
    @staticmethod
    def _is_timeseries_data(data):
        """Check if data contains timeseries structure"""
        return (isinstance(data, dict) and 
                'properties' in data and 
                isinstance(data['properties'], dict) and
                'timeseries' in data['properties'])
    
    @staticmethod
    def _format_timeseries(data, wb, title):
        """Format timeseries data into a structured table"""
        ws = wb.active
        ws.title = "Timeseries Data"
        
        properties = data.get('properties', {})
        timeseries = properties.get('timeseries', [])
        meta = properties.get('meta', {})
        geometry = data.get('geometry', {})
        
        if not timeseries:
            return
        
        # Title section
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        
        # Metadata section
        row = 3
        ws[f'A{row}'] = "Location (Lat, Lon):"
        coords = geometry.get('coordinates', [])
        coord_str = f"{coords[1]}, {coords[0]}" if len(coords) >= 2 else "N/A"
        ws[f'B{row}'] = coord_str
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Updated:"
        ws[f'B{row}'] = meta.get('updated_at', 'N/A')
        ws[f'A{row}'].font = Font(bold=True)
        row += 2
        
        # Parse the first entry to determine column headers
        headers = ['Time']
        sample_entry = timeseries[0] if timeseries else {}
        details_keys = ExcelFormatter._extract_all_detail_keys(timeseries)
        headers.extend(sorted(details_keys))
        
        # Add summary fields if present
        if sample_entry.get('data', {}).get('next_1_hours', {}).get('summary'):
            headers.append('Next 1hr Symbol')
        if sample_entry.get('data', {}).get('next_1_hours', {}).get('details', {}).get('precipitation_amount') is not None:
            headers.append('Next 1hr Precip (mm)')
        
        # Write headers
        header_row = row
        thin_border = Border(
            bottom=Side(style='thin', color='000000')
        )
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        row += 1
        
        # Write data rows
        for entry in timeseries:
            time_val = entry.get('time', '')
            data_section = entry.get('data', {})
            instant_details = data_section.get('instant', {}).get('details', {})
            
            col = 1
            ws.cell(row=row, column=col, value=time_val)
            col += 1
            
            # Write all detail values in order
            for key in sorted(details_keys):
                value = instant_details.get(key, '')
                ws.cell(row=row, column=col, value=value)
                col += 1
            
            # Add next_1_hours summary if present
            next_1hr = data_section.get('next_1_hours', {})
            if 'Next 1hr Symbol' in headers:
                symbol = next_1hr.get('summary', {}).get('symbol_code', '')
                ws.cell(row=row, column=col, value=symbol)
                col += 1
            
            if 'Next 1hr Precip (mm)' in headers:
                precip = next_1hr.get('details', {}).get('precipitation_amount', '')
                ws.cell(row=row, column=col, value=precip)
                col += 1
            
            row += 1
        
        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
        ws.column_dimensions['A'].width = 22  # Time column wider
        
        # Add units sheet if available
        if meta.get('units'):
            units_ws = wb.create_sheet("Units")
            units_ws['A1'] = "Parameter Units"
            units_ws['A1'].font = Font(size=14, bold=True)
            units_ws['A1'].fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            
            units_ws['A2'] = "Parameter"
            units_ws['B2'] = "Unit"
            units_ws['A2'].font = Font(bold=True)
            units_ws['B2'].font = Font(bold=True)
            
            row = 3
            for param, unit in sorted(meta['units'].items()):
                units_ws[f'A{row}'] = param
                units_ws[f'B{row}'] = unit
                row += 1
            
            units_ws.column_dimensions['A'].width = 40
            units_ws.column_dimensions['B'].width = 20
    
    @staticmethod
    def _extract_all_detail_keys(timeseries):
        """Extract all unique detail keys from timeseries"""
        keys = set()
        for entry in timeseries:
            details = entry.get('data', {}).get('instant', {}).get('details', {})
            keys.update(details.keys())
        return keys
    
    @staticmethod
    def _format_generic(data, wb, title):
        """Generic format for non-timeseries data"""
        ws = wb.active
        ws.title = "Response Data"
        
        # Title
        ws['A1'] = title
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        ws.merge_cells('A1:B1')
        
        row = 3
        
        # Metadata section
        if '_metadata' in data:
            ws[f'A{row}'] = "Metadata"
            ws[f'A{row}'].font = Font(size=14, bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            metadata = data['_metadata']
            for key, value in metadata.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = str(value)
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            row += 1
        
        # Error section
        if 'error' in data:
            ws[f'A{row}'] = "Error"
            ws[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            ws[f'A{row}'] = "Error Message"
            ws[f'B{row}'] = data['error']
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            row += 1
        
        # Full data as JSON
        ws[f'A{row}'] = "Full Response (JSON)"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
        ws.merge_cells(f'A{row}:B{row}')
        row += 1
        
        json_str = json.dumps(data, indent=2, ensure_ascii=False)
        ws[f'A{row}'] = json_str
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{row}:B{row + 20}')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 60


if __name__ == "__main__":
    # Test the formatter
    test_data = {
        '_metadata': {
            'api': 'Test API',
            'fetched_at': datetime.now().isoformat()
        },
        'test': 'value',
        'nested': {'key': 'value'}
    }
    
    formatter = ExcelFormatter()
    filename = '/tmp/test_output.xlsx'
    formatter.format(test_data, filename, "Test API Response")
    print(f"Created test Excel file: {filename}")
